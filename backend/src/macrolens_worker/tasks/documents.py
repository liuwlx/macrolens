from __future__ import annotations

import hashlib
import ipaddress
import mimetypes
import re
import socket
from datetime import UTC, datetime
from email.utils import parsedate_to_datetime
from io import BytesIO
from pathlib import PurePosixPath
from urllib.parse import urljoin, urlparse
from uuid import UUID

import fitz  # type: ignore[import-untyped]
import httpx
from lxml import html  # type: ignore[import-untyped]
from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from macrolens_api.config import get_settings
from macrolens_api.models import Document, DocumentChunk, DocumentVersion, Provider, RawObject
from macrolens_api.services.jobs import enqueue_job
from macrolens_api.services.licenses import get_license_for_provider
from macrolens_api.services.storage import ObjectStorage

settings = get_settings()
MAX_DOCUMENT_BYTES = 50 * 1024 * 1024
MAX_REDIRECTS = 5
MAX_EXTRACTED_CHARS = 5_000_000
MAX_PDF_PAGES = 2_000


def _normalized_host(value: str) -> str:
    return value.lower().rstrip(".")


def _allowed_document_domains(provider: Provider) -> set[str]:
    """Return explicit approved hosts, never a registrable-domain approximation.

    Collapsing hosts to the final two labels is unsafe for public suffixes such as co.uk.
    An allowlist entry permits the exact host and its subdomains only. Sibling hosts must be
    declared explicitly in provider metadata.
    """
    domains: set[str] = set()
    for value in (provider.base_url, provider.api_docs_url, provider.terms_url):
        if value:
            host = urlparse(value).hostname
            if host:
                domains.add(_normalized_host(host))
    for value in provider.metadata_json.get("allowed_document_hosts", []):
        host = urlparse(str(value)).hostname or str(value).strip()
        if host:
            domains.add(_normalized_host(host))
    return domains


def _host_allowed(hostname: str, allowed_domains: set[str]) -> bool:
    host = _normalized_host(hostname)
    return any(host == allowed or host.endswith(f".{allowed}") for allowed in allowed_domains)


async def _validate_public_https_url(value: str, *, allowed_domains: set[str] | None = None) -> str:
    parsed = urlparse(value)
    if parsed.scheme != "https" or not parsed.hostname or parsed.username or parsed.password:
        raise RuntimeError("Document URL must be a credential-free HTTPS URL")
    if allowed_domains and not _host_allowed(parsed.hostname, allowed_domains):
        raise RuntimeError("Document URL is outside the selected provider's approved domains")

    def resolve() -> list[str]:
        return list(
            {
                str(item[4][0])
                for item in socket.getaddrinfo(
                    parsed.hostname, parsed.port or 443, type=socket.SOCK_STREAM
                )
            }
        )

    addresses = await __import__("asyncio").to_thread(resolve)
    if not addresses:
        raise RuntimeError("Document host did not resolve")
    for address in addresses:
        ip = ipaddress.ip_address(address)
        if not ip.is_global:
            raise RuntimeError("Document URL resolves to a private or reserved address")
    return value


async def _download_document(
    source_url: str, *, allowed_domains: set[str] | None = None
) -> tuple[bytes, str, str, int, datetime | None]:
    current = await _validate_public_https_url(source_url, allowed_domains=allowed_domains)
    async with httpx.AsyncClient(
        timeout=httpx.Timeout(45.0, connect=15.0),
        headers={"User-Agent": "MacroLens/1.0 (+official-document-ingestion)"},
        follow_redirects=False,
    ) as client:
        for _ in range(MAX_REDIRECTS + 1):
            async with client.stream("GET", current) as response:
                if response.status_code in {301, 302, 303, 307, 308}:
                    location = response.headers.get("location")
                    if not location:
                        raise RuntimeError("Document redirect omitted Location")
                    current = await _validate_public_https_url(
                        urljoin(current, location), allowed_domains=allowed_domains
                    )
                    continue
                response.raise_for_status()
                declared = response.headers.get("content-length")
                if declared:
                    try:
                        if int(declared) > MAX_DOCUMENT_BYTES:
                            raise RuntimeError("Document exceeds the 50 MiB ingestion limit")
                    except ValueError:
                        pass
                body = bytearray()
                async for chunk in response.aiter_bytes():
                    body.extend(chunk)
                    if len(body) > MAX_DOCUMENT_BYTES:
                        raise RuntimeError("Document exceeds the 50 MiB ingestion limit")
                content_type = response.headers.get(
                    "content-type", "application/octet-stream"
                ).split(";", 1)[0]
                last_modified = None
                if response.headers.get("last-modified"):
                    try:
                        last_modified = parsedate_to_datetime(response.headers["last-modified"])
                    except (TypeError, ValueError, OverflowError):
                        last_modified = None
                return bytes(body), content_type, current, response.status_code, last_modified
    raise RuntimeError("Document exceeded redirect limit")


async def fetch_document(
    session: AsyncSession,
    *,
    provider_code: str,
    source_url: str,
    title: str,
    title_zh: str | None = None,
    document_type: str = "official_release",
    external_id: str | None = None,
    language: str = "en",
    published_at: str | None = None,
    copyright_status: str = "official",
    metadata: dict | None = None,
) -> dict[str, str | int]:
    provider = await session.scalar(
        select(Provider).where(Provider.code == provider_code.upper(), Provider.active)
    )
    if provider is None:
        raise RuntimeError(f"Active provider not found: {provider_code}")
    allowed_domains = _allowed_document_domains(provider)
    if not allowed_domains:
        raise RuntimeError(
            "Provider has no approved document domain; configure allowed_document_hosts"
        )
    body, content_type, final_url, http_status, last_modified = await _download_document(
        source_url, allowed_domains=allowed_domains
    )
    suffix = PurePosixPath(urlparse(final_url).path).suffix.lower()
    if not suffix:
        suffix = mimetypes.guess_extension(content_type) or ".bin"
    digest = hashlib.sha256(body).hexdigest()
    date_path = datetime.now(UTC).strftime("%Y/%m/%d")
    key = f"raw/documents/{provider.code.lower()}/{date_path}/{digest}{suffix[:12]}"
    stored = await ObjectStorage().put_bytes(key, body, content_type)
    raw_object = await session.scalar(
        select(RawObject).where(
            RawObject.provider_id == provider.id, RawObject.sha256 == stored.sha256
        )
    )
    if raw_object is None:
        raw_object = RawObject(
            provider_id=provider.id,
            object_uri=stored.uri,
            content_type=stored.content_type,
            byte_size=stored.byte_size,
            sha256=stored.sha256,
            request_url=source_url,
            request_parameters={"final_url": final_url},
            http_status=http_status,
            source_last_modified=last_modified,
        )
        session.add(raw_object)
        await session.flush()
    document = await session.scalar(select(Document).where(Document.source_url == source_url))
    parsed_published_at = None
    if published_at:
        parsed_published_at = datetime.fromisoformat(str(published_at).replace("Z", "+00:00"))
    if document is None:
        document = Document(
            provider_id=provider.id,
            external_id=external_id,
            document_type=document_type,
            title=title,
            title_zh=title_zh,
            source_url=source_url,
            published_at=parsed_published_at,
            language=language,
            copyright_status=copyright_status,
            status="processing",
            metadata_json={**(metadata or {}), "final_url": final_url},
        )
        session.add(document)
        await session.flush()
    else:
        if document.provider_id != provider.id:
            raise RuntimeError("Document URL is already registered under another provider")
        document.title = title
        document.title_zh = title_zh
        document.document_type = document_type
        document.external_id = external_id or document.external_id
        document.language = language
        document.published_at = parsed_published_at or document.published_at
        document.copyright_status = copyright_status
        document.status = "processing"
        document.metadata_json = {
            **document.metadata_json,
            **(metadata or {}),
            "final_url": final_url,
        }
    await enqueue_job(
        session,
        job_type="parse_document",
        payload={"document_id": str(document.id), "raw_object_id": str(raw_object.id)},
        idempotency_key=f"parse-document:{document.id}:{raw_object.sha256}",
        priority=10,
        max_attempts=3,
    )
    await session.commit()
    return {
        "status": "queued_for_parsing",
        "document_id": str(document.id),
        "raw_object_id": str(raw_object.id),
        "byte_size": stored.byte_size,
    }


def _chunk_text(text: str, max_chars: int = 3500, overlap: int = 300) -> list[str]:
    if max_chars <= 0:
        raise ValueError("max_chars must be positive")
    overlap = max(0, min(overlap, max_chars - 1))
    normalized = re.sub(r"\n{3,}", "\n\n", text).strip()
    if not normalized:
        return []

    chunks: list[str] = []
    start = 0
    length = len(normalized)
    while start < length:
        end = min(start + max_chars, length)
        if end < length:
            # Prefer a natural boundary in the latter half of the window, while retaining
            # deterministic maximum size for OCR/minified sources with no whitespace.
            lower = start + max_chars // 2
            candidates = [
                normalized.rfind("\n\n", lower, end),
                normalized.rfind("\n", lower, end),
                normalized.rfind(" ", lower, end),
            ]
            boundary = max(candidates)
            if boundary > start:
                end = boundary
        if end <= start:
            end = min(start + max_chars, length)
        chunks.append(normalized[start:end])
        if end >= length:
            break
        start = max(end - overlap, start + 1)
    return chunks


def _bounded_text(value: str) -> str:
    if len(value) > MAX_EXTRACTED_CHARS:
        raise RuntimeError("Extracted document text exceeds the safety limit")
    return value


def _extract_pdf(raw: bytes) -> tuple[str, list[tuple[int, str]]]:
    document = fitz.open(stream=raw, filetype="pdf")
    if document.page_count > MAX_PDF_PAGES:
        raise RuntimeError("PDF exceeds the page-count safety limit")
    pages: list[tuple[int, str]] = []
    total = 0
    for index, page in enumerate(document, start=1):
        text = page.get_text("text")
        total += len(text)
        if total > MAX_EXTRACTED_CHARS:
            raise RuntimeError("Extracted document text exceeds the safety limit")
        pages.append((index, text))
    return "\n\n".join(text for _page, text in pages), pages


def _extract_html(raw: bytes) -> str:
    document = html.fromstring(raw)
    for element in document.xpath("//script|//style|//nav|//footer"):
        element.drop_tree()
    return _bounded_text(
        "\n".join(line.strip() for line in document.text_content().splitlines() if line.strip())
    )


def _extract_xlsx(raw: bytes) -> str:
    workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    lines: list[str] = []
    total = 0
    for worksheet in workbook.worksheets:
        lines.append(f"# {worksheet.title}")
        for row in worksheet.iter_rows(values_only=True):
            text = "\t".join("" if value is None else str(value) for value in row).rstrip()
            if not text:
                continue
            total += len(text) + 1
            if total > MAX_EXTRACTED_CHARS:
                raise RuntimeError("Extracted spreadsheet text exceeds the safety limit")
            lines.append(text)
    return "\n".join(lines)


async def parse_document(
    session: AsyncSession,
    *,
    document_id: UUID,
    raw_object_id: UUID,
) -> dict[str, int | str]:
    document = await session.get(Document, document_id)
    raw_object = await session.get(RawObject, raw_object_id)
    if document is None or raw_object is None:
        raise RuntimeError("Document or raw object not found")
    key = raw_object.object_uri.split(f"s3://{settings.s3_bucket}/", 1)[-1]
    raw = await ObjectStorage().get_bytes(key)
    content_type = raw_object.content_type or ""
    page_chunks: list[tuple[int | None, int | None, str]] = []
    lower_url = document.source_url.lower()
    if "pdf" in content_type or lower_url.endswith(".pdf"):
        extracted, pages = _extract_pdf(raw)
        for page_no, page_text in pages:
            for chunk in _chunk_text(page_text):
                page_chunks.append((page_no, page_no, chunk))
    elif "html" in content_type or lower_url.endswith((".html", ".htm")):
        extracted = _extract_html(raw)
        page_chunks = [(None, None, chunk) for chunk in _chunk_text(extracted)]
    elif "spreadsheetml" in content_type or lower_url.endswith(".xlsx"):
        extracted = _extract_xlsx(raw)
        page_chunks = [(None, None, chunk) for chunk in _chunk_text(extracted)]
    elif content_type.startswith("text/") or lower_url.endswith(
        (".txt", ".md", ".csv", ".xml", ".json")
    ):
        extracted = _bounded_text(raw.decode("utf-8", errors="replace"))
        page_chunks = [(None, None, chunk) for chunk in _chunk_text(extracted)]
    else:
        raise RuntimeError(f"Unsupported document content type: {content_type or 'unknown'}")
    if not extracted.strip():
        raise RuntimeError(
            "Document extraction produced no text; refusing to mark an incomplete document active"
        )
    if not page_chunks:
        raise RuntimeError(
            "Document extraction produced no searchable chunks; refusing incomplete publication"
        )
    digest = hashlib.sha256(extracted.encode("utf-8")).hexdigest()
    existing = await session.scalar(
        select(DocumentVersion).where(
            DocumentVersion.document_id == document_id,
            DocumentVersion.content_hash == digest,
        )
    )
    if existing:
        document.status = "active"
        await session.commit()
        return {"status": "unchanged", "chunks": 0, "version_id": str(existing.id)}
    latest = await session.scalar(
        select(DocumentVersion)
        .where(DocumentVersion.document_id == document_id)
        .order_by(DocumentVersion.version_no.desc())
        .limit(1)
    )
    if latest:
        latest.superseded_at = datetime.now(UTC)
    version = DocumentVersion(
        document_id=document_id,
        version_no=(latest.version_no + 1) if latest else 1,
        content_hash=digest,
        raw_object_id=raw_object_id,
        extracted_text=extracted,
        effective_at=datetime.now(UTC),
        parser_version="macrolens-parser-v1",
    )
    session.add(version)
    await session.flush()
    for index, (page_start, page_end, content) in enumerate(page_chunks):
        session.add(
            DocumentChunk(
                document_version_id=version.id,
                chunk_no=index,
                page_start=page_start,
                page_end=page_end,
                content=content,
                token_count=max(1, len(content) // 4),
            )
        )
    document.status = "active"
    await session.commit()
    if settings.openai_api_key:
        await enqueue_job(
            session,
            job_type="embed_document",
            payload={"document_version_id": str(version.id)},
            idempotency_key=f"embed-document:{version.id}:{settings.openai_embedding_model}",
            priority=8,
            max_attempts=3,
        )
        await enqueue_job(
            session,
            job_type="summarize_document",
            payload={"document_version_id": str(version.id)},
            idempotency_key=f"summarize-document:{version.id}:{settings.openai_model}",
            priority=6,
            max_attempts=3,
        )
    return {"status": "parsed", "chunks": len(page_chunks), "version_id": str(version.id)}


async def embed_document(
    session: AsyncSession, *, document_version_id: UUID
) -> dict[str, int | str]:
    if not settings.openai_api_key:
        raise RuntimeError("OPENAI_API_KEY is required for embeddings")
    version = await session.get(DocumentVersion, document_version_id)
    if version is None:
        raise RuntimeError("Document version not found")
    document = await session.get(Document, version.document_id)
    if document is None:
        raise RuntimeError("Document not found")
    license_info = await get_license_for_provider(session, document.provider_id)
    if not license_info.ai_context_allowed:
        raise RuntimeError("Document license does not allow AI context or embeddings")
    chunks = list(
        (
            await session.scalars(
                select(DocumentChunk)
                .where(
                    DocumentChunk.document_version_id == document_version_id,
                    DocumentChunk.embedding.is_(None),
                )
                .order_by(DocumentChunk.chunk_no)
            )
        ).all()
    )
    if not chunks:
        return {"status": "unchanged", "embedded": 0}
    from openai import AsyncOpenAI

    client = AsyncOpenAI(api_key=settings.openai_api_key, base_url=settings.openai_base_url)
    embedded = 0
    for start in range(0, len(chunks), 64):
        batch = chunks[start : start + 64]
        response = await client.embeddings.create(
            model=settings.openai_embedding_model,
            input=[chunk.content for chunk in batch],
        )
        for chunk, item in zip(batch, response.data, strict=True):
            chunk.embedding = item.embedding
            chunk.embedding_model = settings.openai_embedding_model
            embedded += 1
    await session.commit()
    return {"status": "embedded", "embedded": embedded}


async def summarize_document(session: AsyncSession, *, document_version_id: UUID) -> dict[str, str]:
    if not settings.openai_api_key:
        raise RuntimeError("OPENAI_API_KEY is required for document summaries")
    version = await session.get(DocumentVersion, document_version_id)
    if version is None:
        raise RuntimeError("Document version not found")
    document = await session.get(Document, version.document_id)
    if document is None:
        raise RuntimeError("Document not found")
    license_info = await get_license_for_provider(session, document.provider_id)
    if not license_info.ai_context_allowed:
        raise RuntimeError("Document license does not allow AI summarization")
    if version.ai_summary_zh:
        return {"status": "unchanged", "version_id": str(version.id)}
    source = (version.extracted_text or "").strip()
    if not source:
        raise RuntimeError("Document has no extracted text")
    from openai import AsyncOpenAI

    client = AsyncOpenAI(api_key=settings.openai_api_key, base_url=settings.openai_base_url)
    response = await client.responses.create(
        model=settings.openai_model,
        input=[
            {
                "role": "system",
                "content": (
                    "你是宏观研究文档编辑。仅基于原文生成中文摘要，包含：核心结论、关键数据、"
                    "口径与限制。不得补充原文没有的信息。控制在500字以内。"
                ),
            },
            {"role": "user", "content": source[:120_000]},
        ],
        store=settings.openai_store,
    )
    version.ai_summary_zh = response.output_text.strip()
    await session.commit()
    return {"status": "summarized", "version_id": str(version.id)}
